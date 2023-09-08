# import statements
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from user_settings import *
from application.constants import *
import json


# enable/disable debug
print_metadata = False   # if True, script will print metadata script

# function for filter logic
def device_filter(device):
    # if filter is not set, always return true
    if device_list is None:
        return True
    # if device_list is a string, turn it into list
    if isinstance(device_list, str):
        local_filter = [device_list]
    elif isinstance(device_list, list):
        local_filter = device_list
    else:
        local_filter = None
        print("Error: 'device_list' needs to be either a string or a list.")
        quit(1)
    # filter devices
    if device in local_filter:
        result = True
    else:
        result = False
    # return result
    return result if negate is False else not result


# Extract data from Excel file into dictionary object
# # Load the Excel file
workbook = load_workbook(file_path, data_only=True)

# # Select the active worksheet
if active_sheet is None:
    worksheet = workbook.active
else:
    worksheet = workbook[active_sheet]

# # define skipable values
skipable = [None, '', 'N/A']

# # set host/vdom columns and check for duplicate entries
duplicate_check = []
host_column = None
vdom_column = None
for column in worksheet.iter_cols():
    # assign friendly names
    header = column[0].value
    current_column = column[0].column
    # check for hostname column
    if header == HOSTNAME:
        host_column = current_column
    # check for vdom column
    if header == VDOM:
        vdom_column = current_column
    # add value to duplicate_check list
    if header not in skipable:
        if header in duplicate_check:
            print(f"Error: variable \"{header}\" is duplicated at column {get_column_letter(current_column)}.")
            quit(2)
        duplicate_check.append(header)
if host_column is None:
    print(f"Error: {HOSTNAME} column is not present.")
    quit(101)
if vdom_column is None:
    print(f"Error: {VDOM} column is not present.")
    quit(102)
start_column = max([host_column, vdom_column]) + 1

# # create a list to hold variable values
variables = []

# # define data for each variable
for column in worksheet.iter_cols(min_col=start_column):
    if column[0].value not in skipable:
        # define variable inner-dictionary
        variable = {
            "name": column[0].value
        }
        global_value = column[1].value
        mapping = []

        # define mapping for each variable
        for cell in column[2:]:
            host = worksheet[get_column_letter(host_column) + str(cell.row)].value
            vdom = worksheet[get_column_letter(vdom_column) + str(cell.row)].value
            if cell.value not in skipable and device_filter(host):
                scope = [{
                    DEVICE: host
                }]
                if vdom not in skipable and vdom != GLOBAL:
                    scope[0][VIRTUAL_DOMAIN] = vdom
                device_settings = {
                    SCOPE: scope,
                    VALUE: str(cell.value)
                }
                mapping.append(device_settings)

        # fill in variable dictionary
        if mapping:  # if there are device mappings to be had, add them to the variable dictionary
            variable[MAPPING] = mapping
        if global_value not in skipable and update_defaults:  # if a global value is present and the update boolean is set, include a global value
            variable[VALUE] = str(global_value)
        if VALUE in variable or MAPPING in variable:  # if a variable has data, add it to the template
            variables.append(variable)

# Convert to JSON
json_str = json.dumps(variables, indent=2)
if print_metadata:
    print(json_str)

# Save output to file 'output.json'
with open(f"{OUTPUT}", 'w') as f:
    f.write(json_str)
